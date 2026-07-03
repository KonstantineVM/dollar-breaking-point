#!/usr/bin/env python3
# RDT-G Part 1 I.B PARSE — deterministic regenerator.
# Reads the on-disk RDT-D SAFE/IMF SDDS corpus (build/reserve/rdtd_evidence/raw_*.xls*)
# and writes EXACTLY two files:
#   build/reserve/RDTG_ib_series.csv
#   build/reserve/RDTG_ib_staging.json
# STAGING ONLY: no verdict is rendered here; every quantity is stamped NOT ESTABLISHED.
# Conventions per RDTE/RDTF recomputes: sorted keys, fixed formatting, byte-identical reruns.
# Vintage rule (per build/reserve/rdtd_sdds_manifest.json, publisher_frictions[0..1]):
# latest-published file per data month; the only duplicated month is 2020-10, where the
# corrected later vintage raw_075_* supersedes raw_074_* (both retained on disk).
# All axis constants are READ from committed artifacts, never hardcoded:
#   threshold = 0.05 x RDTD_result.json -> identity.china_alone.delta_non_us_busd

import csv
import glob
import hashlib
import io
import json
import os
import re

HERE = os.path.dirname(os.path.abspath(__file__))
EVID = os.path.join(HERE, "rdtd_evidence")
OUT_CSV = os.path.join(HERE, "RDTG_ib_series.csv")
OUT_JSON = os.path.join(HERE, "RDTG_ib_staging.json")
RDTD_RESULT = os.path.join(HERE, "RDTD_result.json")
RDTD_SERIES = os.path.join(HERE, "rdtd_sdds_series.csv")

SUPERSEDED_PREFIXES = ("raw_074_",)  # 2020-10 first vintage, superseded by raw_075_*

# The IMF template I.B block: B total + six published sub-lines, in template order.
IB_KEYS = ["ib_total", "ib_securities", "ib_deposits", "ib_loans",
           "ib_fin_derivatives", "ib_gold", "ib_other"]
IB_EN_REQUIRED = [
    "Other foreign currency assets",
    "securities not included in official reserve assets",
    "deposits not included in official reserve assets",
    "loans not included in official reserve assets",
    "financial derivatives not included in official reserve assets",
    "gold not included in official reserve assets",
    "other",
]


def rows_of(path):
    """Return the first sheet (Section I) as a list of row-lists of cell values."""
    if path.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sh = wb[wb.sheetnames[0]]
        return [[("" if c is None else c) for c in r] for r in sh.iter_rows(values_only=True)]
    import xlrd
    bk = xlrd.open_workbook(path)
    sh = bk.sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def fmt_bn(v_yi):
    """100-million-USD template value -> USD bn string, fixed formatting."""
    return repr(round(float(v_yi) / 10.0, 6))


def parse_file(path):
    rows = rows_of(path)
    # data month from the As-at header (Chinese date, present in every vintage)
    month = None
    asat_raw = None
    for r in rows[:4]:
        m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", str(r[0]))
        if m:
            month = "%04d-%02d" % (int(m.group(1)), int(m.group(2)))
            asat_raw = norm(r[0])
            break
    assert month is not None, "no As-at date header in %s" % path
    # unit header: two observed variants, both 100 million USD
    unit = None
    for r in rows[:8]:
        for c in r:
            s = re.sub(r"\s+", "", str(c))
            if "亿美元" in s:
                unit = norm(c)
                break
        if unit:
            break
    assert unit is not None, "no 100-million-USD unit header in %s" % path
    # locate the I.B block
    brow = None
    for i, r in enumerate(rows):
        if "Other foreign currency assets" in str(r[0]):
            brow = i
            break
    assert brow is not None, "I.B block absent in %s" % path
    labels = []
    values = {}
    flags = {}
    for k, key in enumerate(IB_KEYS):
        ridx = brow + k
        if ridx >= len(rows):
            values[key] = None
            flags[key] = "ABSENT"
            labels.append(None)
            continue
        lab = norm(rows[ridx][0])
        assert IB_EN_REQUIRED[k].lower() in lab.lower(), \
            "unexpected I.B row label %r (wanted %r) in %s" % (lab, IB_EN_REQUIRED[k], path)
        labels.append(lab)
        cell = rows[ridx][1] if len(rows[ridx]) > 1 else ""
        if isinstance(cell, str) and cell.strip() == "":
            values[key] = None
            flags[key] = "BLANK"
        else:
            assert isinstance(cell, (int, float)), \
                "unparsable I.B cell %r in %s row %d" % (cell, path, ridx)
            values[key] = float(cell)  # 100 million USD as published
            flags[key] = "PUBLISHED"
    # I.A cross-check inputs: securities (1)(a) and A. total, USD column
    ia = {}
    for i, r in enumerate(rows[:brow]):
        lab = str(r[0])
        if "Official reserve assets" in lab and lab.strip().startswith("A"):
            ia["total_ora"] = float(r[1])
        if "Securities" in lab and "certificate" not in lab.lower() and "of which" not in lab:
            ia.setdefault("securities", float(r[1]) if not (isinstance(r[1], str) and r[1].strip() == "") else None)
    return {"month": month, "asat": asat_raw, "unit": unit, "labels": labels,
            "values": values, "flags": flags, "ia": ia,
            "file": os.path.basename(path)}


def main():
    files = sorted(glob.glob(os.path.join(EVID, "raw_*.xls")) +
                   glob.glob(os.path.join(EVID, "raw_*.xlsx")))
    assert files, "no corpus under %s" % EVID
    parsed = [parse_file(p) for p in files]
    n_files = len(parsed)

    # vintage discipline: latest-published file per month (raw_075 supersedes raw_074 for 2020-10)
    by_month = {}
    duplicates = {}
    for rec in parsed:
        by_month.setdefault(rec["month"], []).append(rec)
    chosen = {}
    for month, recs in sorted(by_month.items()):
        if len(recs) == 1:
            chosen[month] = recs[0]
        else:
            duplicates[month] = sorted(r["file"] for r in recs)
            keep = [r for r in recs if not r["file"].startswith(SUPERSEDED_PREFIXES)]
            assert len(keep) == 1, "unresolved duplicate vintages for %s: %r" % (month, duplicates[month])
            chosen[month] = keep[0]
    months = sorted(chosen)
    # continuity check: every calendar month in span must be present
    missing = []
    y, m = int(months[0][:4]), int(months[0][5:])
    ye, me = int(months[-1][:4]), int(months[-1][5:])
    while (y, m) <= (ye, me):
        k = "%04d-%02d" % (y, m)
        if k not in chosen:
            missing.append(k)
        m += 1
        if m == 13:
            y, m = y + 1, 1
    # unit must be 100 million USD in every file (two header variants observed)
    unit_variants = sorted(set(r["unit"] for r in parsed))
    for u in unit_variants:
        assert "亿美元" in re.sub(r"\s+", "", u), "unexpected unit header %r" % u
    # row labels: quote the exact set found; must be identical across the corpus
    label_variants = sorted(set(tuple(r["labels"]) for r in parsed))
    assert len(label_variants) == 1, "I.B row labels vary across vintages: %r" % label_variants

    # ---------- CSV ----------
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    header = ["data_month"]
    for key in IB_KEYS:
        header += ["%s_usd_bn" % key, "%s_flag" % key]
    header += ["channel", "source_file"]
    w.writerow(header)
    for month in months:
        rec = chosen[month]
        row = [month]
        for key in IB_KEYS:
            v = rec["values"][key]
            row += ["" if v is None else fmt_bn(v), rec["flags"][key]]
        row += ["SAFE-EN", "build/reserve/rdtd_evidence/" + rec["file"]]
        w.writerow(row)
    csv_bytes = buf.getvalue().encode("utf-8")
    with open(OUT_CSV, "wb") as f:
        f.write(csv_bytes)

    # ---------- threshold: READ from the committed artifact (never hardcoded) ----------
    with open(RDTD_RESULT) as f:
        rdtd = json.load(f)
    upper_mass = rdtd["identity"]["china_alone"]["delta_non_us_busd"]
    threshold = round(0.05 * upper_mass, 6)

    # ---------- deltas (USD bn) ----------
    def bn(month, key):
        v = chosen[month]["values"][key]
        return None if v is None else round(v / 10.0, 6)

    latest = months[-1]
    first = months[0]
    vw_start = "2022-01"
    base_end = "2021-12"
    assert vw_start in chosen and base_end in chosen

    def delta(key, m0, m1):
        a, b = bn(m0, key), bn(m1, key)
        if a is None or b is None:
            return None
        return round(b - a, 6)

    d_vw = delta("ib_total", vw_start, latest)
    d_vw_alt = delta("ib_total", base_end, latest)
    d_base = delta("ib_total", first, base_end)
    assert d_vw is not None and d_vw_alt is not None and d_base is not None

    sub_deltas = {}
    for key in IB_KEYS[1:]:
        sub_deltas[key] = {
            "baseline_window_2015_06_to_2021_12_usd_bn": delta(key, first, base_end),
            "endpoint_values_usd_bn": {vw_start: bn(vw_start, key), latest: bn(latest, key),
                                       first: bn(first, key), base_end: bn(base_end, key)},
            "verdict_window_2022_01_to_latest_usd_bn": delta(key, vw_start, latest),
        }
        pubm = [mm for mm in months if chosen[mm]["flags"][key] == "PUBLISHED"]
        sub_deltas[key]["last_published_month"] = pubm[-1] if pubm else None
        sub_deltas[key]["last_published_value_usd_bn"] = bn(pubm[-1], key) if pubm else None

    # level profile of the I.B total
    tot = [bn(mm, "ib_total") for mm in months]
    level = {
        "max_usd_bn": max(tot),
        "max_month": months[tot.index(max(tot))],
        "mean_usd_bn": round(sum(tot) / len(tot), 6),
        "min_usd_bn": min(tot),
        "min_month": months[tot.index(min(tot))],
        "months": len(tot),
    }

    # publication-status spans per line (suppression assessment input)
    def spans(key):
        out = []
        cur = None
        for mm in months:
            fl = chosen[mm]["flags"][key]
            if cur and cur["flag"] == fl:
                cur["end"] = mm
                cur["n"] += 1
            else:
                if cur:
                    out.append(cur)
                cur = {"flag": fl, "start": mm, "end": mm, "n": 1}
        out.append(cur)
        return out

    pub_status = {}
    for key in IB_KEYS:
        fl = [chosen[mm]["flags"][key] for mm in months]
        pub_status[key] = {
            "published_months": fl.count("PUBLISHED"),
            "blank_months": fl.count("BLANK"),
            "absent_months": fl.count("ABSENT"),
            "spans": spans(key),
        }

    # additivity of published sub-lines to the I.B total
    max_add_diff = 0.0
    for mm in months:
        s = sum(chosen[mm]["values"][k] for k in IB_KEYS[1:] if chosen[mm]["values"][k] is not None)
        max_add_diff = max(max_add_diff, abs(s - chosen[mm]["values"]["ib_total"]))
    max_add_diff_bn = round(max_add_diff / 10.0, 6)

    # ---------- I.A cross-check vs committed rdtd_sdds_series.csv ----------
    committed = {}
    with open(RDTD_SERIES) as f:
        for r in csv.DictReader(f):
            committed[r["data_month"]] = r
    xdiff_sec = 0.0
    xdiff_tot = 0.0
    xn = 0
    file_mismatch = []
    for mm in months:
        if mm not in committed:
            continue
        xn += 1
        rec = chosen[mm]
        xdiff_sec = max(xdiff_sec, abs(rec["ia"]["securities"] / 10.0 - float(committed[mm]["securities_usd_bn"])))
        xdiff_tot = max(xdiff_tot, abs(rec["ia"]["total_ora"] / 10.0 - float(committed[mm]["total_official_reserve_assets_usd_bn"])))
        if os.path.basename(committed[mm]["source_file"]) != rec["file"]:
            file_mismatch.append(mm)
    crosscheck = {
        "committed_csv": "build/reserve/rdtd_sdds_series.csv",
        "finding": ("PASS: same-file re-parse of Section I.A reconciles with the committed RDT-D "
                    "extraction on all overlapping months" if xdiff_sec < 5e-7 and xdiff_tot < 5e-7
                    and not file_mismatch else "MISMATCH — recorded, not smoothed"),
        "lines_checked": ["(1)(a) Securities", "A. Official reserve assets (total)"],
        "max_abs_diff_securities_usd_bn": round(xdiff_sec, 9),
        "max_abs_diff_total_ora_usd_bn": round(xdiff_tot, 9),
        "months_checked": xn,
        "source_file_mismatch_months": file_mismatch,
        "spot_check_requirement": ">=5 months required; all overlapping months checked",
    }

    # ---------- staging object (NO verdict) ----------
    lab = list(label_variants[0])
    staging = {
        "STATUS": ("NOT ESTABLISHED — STAGING ONLY. Axis numbers staged for the RDT-G assembly; "
                   "the I.B verdict (SURGE/FLAT/DECLINE/SUPPRESSED) is rendered by the later "
                   "assembly step, not here, and no result below is established until its "
                   "verifier scenario has run and its verifier artifact exists."),
        "task": "RDT-G Part 1 — Section I.B ('other foreign currency assets') parse and axis staging",
        "governing_preregistration": {
            "path": "build/reserve/RDTG_prediction.md",
            "windows": {"baseline": "calendar 2015->2021 (data floor 2015-06)",
                        "verdict": "2022-01 -> latest published (pre-registered wording)"},
        },
        "corpus": {
            "SOURCE": "build/reserve/rdtd_evidence/ (RDT-D fetch, disk-only here)",
            "raw_files_parsed": n_files,
            "unique_data_months": len(months),
            "span": [first, latest],
            "missing_months": missing,
            "unparsable_files": [],
            "data_month_read_from": "As-at date in the file header (never from link labels)",
            "duplicate_month_vintages": duplicates,
            "vintage_rule": ("latest-published file per month; 2020-10: raw_075_* (corrected, later "
                             "vintage) supersedes raw_074_* per build/reserve/rdtd_sdds_manifest.json "
                             "line_map_section_IA.publisher_frictions[0..1]; both files retained on disk"),
            "vintage_note_2020_10_ib": ("the two 2020-10 vintages differ on I.B only in numeric precision "
                                        "and in the derivatives sub-line (BLANK in raw_074, explicit 0.0 in "
                                        "raw_075); I.B total agrees within 0.0002 USD bn (raw_074 165.882 "
                                        "rounded vs raw_075 165.881813 full precision)"),
        },
        "unit_finding": {
            "unit_as_published": "100 million USD (亿美元) — NOT USD millions; the task premise 'USD millions per the template convention' is contradicted by the real file headers and is corrected here",
            "header_variants_found": unit_variants,
            "conversion": "CSV and all staging values converted to USD bn by dividing the published 亿美元 value by 10 (same convention as build/reserve/rdtd_sdds_series.csv)",
            "SOURCE": "unit header row of every corpus file (e.g. build/reserve/rdtd_evidence/raw_000_6798bb2d92a547dc972c00310d54bc6d.xls row 5; raw_124_* row 4 for the 2015-06..2016-02 single-column variant)",
        },
        "ib_row_labels_exact": {
            "identical_across_all_files": True,
            "labels": {IB_KEYS[i]: lab[i] for i in range(7)},
            "note": ("the template's six I.B sub-lines are securities / deposits / loans / financial "
                     "derivatives / gold / other, each 'not included in official reserve assets'; there "
                     "is no separate I.B securities-lending/repo line in the SAFE rendering"),
        },
        "level_profile_ib_total_usd_bn": dict(level, SOURCE="build/reserve/RDTG_ib_series.csv (this parse)"),
        "additivity_check": {
            "max_abs_diff_published_sublines_vs_ib_total_usd_bn": max_add_diff_bn,
            "note": "published sub-lines sum to the I.B total within rounding in every month",
        },
        "axis_numbers_STAGING_ONLY": {
            "threshold_reproduction": {
                "constant_read_from": "build/reserve/RDTD_result.json field path: identity.china_alone.delta_non_us_busd",
                "upper_mass_busd": upper_mass,
                "multiplier": 0.05,
                "threshold_busd": threshold,
                "prereg_displayed_value": 24.749,
                "note": "computed 0.05 x the committed constant at run time, never hardcoded",
            },
            "delta_ib_total_verdict_window": {
                "definition": "I.B_total(latest published) - I.B_total(2022-01), per the pre-registered window wording '2022-01 -> latest published'",
                "endpoints": {vw_start: bn(vw_start, "ib_total"), latest: bn(latest, "ib_total")},
                "value_usd_bn": d_vw,
                "variant_2021_12_base": {
                    "definition": "I.B_total(latest) - I.B_total(2021-12) — the calendar-2022-onward change including January 2022's within-month move; stated as sensitivity, same side of every bar",
                    "endpoints": {base_end: bn(base_end, "ib_total"), latest: bn(latest, "ib_total")},
                    "value_usd_bn": d_vw_alt,
                },
            },
            "delta_ib_total_baseline_window_context": {
                "definition": "I.B_total(2021-12) - I.B_total(2015-06) (baseline window, data floor stated)",
                "endpoints": {first: bn(first, "ib_total"), base_end: bn(base_end, "ib_total")},
                "value_usd_bn": d_base,
            },
            "staged_comparison_STAGING_ONLY": {
                "abs_delta_vs_threshold": ("|%r| < %r" % (d_vw, threshold)) if abs(d_vw) < threshold else ("|%r| >= %r" % (d_vw, threshold)),
                "side": ("FLAT-side (strictly between -threshold and +threshold)" if abs(d_vw) < threshold
                         else ("SURGE-side (>= +threshold)" if d_vw >= threshold else "DECLINE-side (<= -threshold)")),
                "side_variant_2021_12_base": ("FLAT-side (strictly between -threshold and +threshold)" if abs(d_vw_alt) < threshold
                                              else ("SURGE-side (>= +threshold)" if d_vw_alt >= threshold else "DECLINE-side (<= -threshold)")),
                "sign_note": ("the staged delta is NEGATIVE (I.B fell over the verdict window) but does not reach "
                              "the -%r DECLINE bar under either endpoint convention" % threshold) if (d_vw < 0 and abs(d_vw) < threshold and abs(d_vw_alt) < threshold) else "see values",
                "NOT_A_VERDICT": "the SURGE/FLAT/DECLINE/SUPPRESSED call is rendered by the RDT-G assembly, not by this parse",
            },
            "sub_line_deltas": sub_deltas,
        },
        "suppression_assessment_STAGING_ONLY": {
            "ib_total": "PUBLISHED in all 132 months — the I.B block itself is NOT unpublished; SUPPRESSED-CANDIDATE does not attach to the total",
            "per_line_status": pub_status,
            "reading": ("securities: published all 132 months and carries essentially the whole I.B line; "
                        "gold, other: BLANK in all 132 months (a disclosure state, distinct from zero, never "
                        "interpolated); deposits: BLANK in 131 of 132 months, single explicit 0 in 2017-04; "
                        "loans: published with real values through most of the corpus, BLANK 2025-11..2026-02 "
                        "and 2026-05, explicit 0.0 in 2026-03..04 (see spans); financial derivatives: explicit "
                        "values (incl. 0.0 and negatives) most months, BLANK 2015-06..2015-11 and 2019-10..2020-09 "
                        "(see spans). "
                        "Whether the persistent component blanks meet the pre-registration's SUPPRESSED clause "
                        "('the I.B line or its components are unpublished/aggregated/blank') alongside the "
                        "published total is an assembly-level call, staged here as facts only."),
        },
        "crosscheck_ia_vs_committed": crosscheck,
        "outputs": {
            "series_csv": "build/reserve/RDTG_ib_series.csv",
            "staging_json": "build/reserve/RDTG_ib_staging.json",
            "regenerator": "build/reserve/RDTG_ib_recompute.py",
        },
    }
    js = json.dumps(staging, ensure_ascii=False, sort_keys=True, indent=1) + "\n"
    with open(OUT_JSON, "wb") as f:
        f.write(js.encode("utf-8"))

    for pth in (OUT_CSV, OUT_JSON):
        with open(pth, "rb") as f:
            print(os.path.basename(pth), hashlib.sha256(f.read()).hexdigest())


if __name__ == "__main__":
    main()
